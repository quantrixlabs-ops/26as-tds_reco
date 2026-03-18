"""
Batch Multi-Party Excel Generator
Produces a single combined workbook with:
  - Sheet 1: Master Summary (one row per party)
  - Per party: {Name}_Matched, {Name}_Unmatched, {Name}_Variance
"""
from __future__ import annotations

import io
import re
import statistics
from datetime import datetime
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import DEFAULT_FINANCIAL_YEAR, VARIANCE_CAP_FORCE_SINGLE as VARIANCE_CAP_PCT
from models import CleaningReport, MatchedPair, RecoResult, UnmatchedAs26Entry

# Import styling helpers from the single-party generator
from excel_generator import (
    NAVY, WHITE, LIGHT_BLUE,
    VAR_GREEN, VAR_YELLOW, VAR_RED,
    CONF_HIGH, CONF_MEDIUM, CONF_LOW,
    VIOLATION_RED, GOOD_GREEN,
    INR_FMT,
    _fill, _font, _border, _align,
    _header_style, _col_header, _data_cell, _autofit,
    _var_color, _conf_color,
    _build_matched, _build_unmatched_26as, _build_unmatched_books, _build_variance,
)


def _safe_sheet_name(name: str, suffix: str, used_names: Set[str]) -> str:
    """
    Create an Excel-safe sheet name (max 31 chars, no special chars).
    Truncate party name and append suffix, deduplicate with numeric counter.
    """
    # Clean name: remove chars illegal in Excel sheet names
    clean = re.sub(r"[\\/*?\[\]:]", "", name).strip()
    # Max length for name part: 31 - len(suffix) - 1 (for underscore)
    max_name_len = 31 - len(suffix) - 1
    if max_name_len < 5:
        max_name_len = 5
    clean = clean[:max_name_len]
    candidate = f"{clean}_{suffix}"

    # Deduplicate
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    for i in range(2, 100):
        alt = f"{clean[:max_name_len - len(str(i)) - 1]}{i}_{suffix}"
        if alt not in used_names:
            used_names.add(alt)
            return alt

    # Fallback
    fallback = f"Party{len(used_names)}_{suffix}"
    used_names.add(fallback)
    return fallback


def _build_master_summary(
    ws,
    results: List[Tuple[RecoResult, CleaningReport, str]],
    fy_label: str,
):
    """Build the Master Summary sheet with one row per party."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    now_str = datetime.now().strftime("%d-%b-%Y %H:%M")
    title = f"Batch TDS Reconciliation | {fy_label} | {len(results)} Parties | Generated: {now_str}"
    _header_style(ws, 1, 1, 16, title)

    headers = [
        "#", "Deductor Name", "TAN", "SAP File",
        "Match Rate %", "Matched", "Total 26AS", "Unmatched 26AS",
        "Unmatched Books", "Violations",
        "HIGH", "MEDIUM", "LOW", "Cross-FY",
        "Avg Var %", "Status",
    ]
    for c, h in enumerate(headers, 1):
        _col_header(ws, 2, c, h)

    # Aggregate totals
    total_26as = 0
    total_matched = 0
    total_unmatched_26as = 0
    total_unmatched_books = 0
    total_violations = 0
    total_high = 0
    total_med = 0
    total_low = 0

    for r, (result, report, sap_fn) in enumerate(results, 3):
        match_bg = (
            VAR_GREEN if result.match_rate_pct >= 95 else
            VAR_YELLOW if result.match_rate_pct >= 75 else VAR_RED
        )
        viol_bg = VIOLATION_RED if result.constraint_violations > 0 else None

        _data_cell(ws, r, 1,  r - 2, align_h="center")
        _data_cell(ws, r, 2,  result.deductor_name, bold=True)
        _data_cell(ws, r, 3,  result.tan, align_h="center")
        _data_cell(ws, r, 4,  sap_fn)
        _data_cell(ws, r, 5,  f"{result.match_rate_pct:.1f}%", bg=match_bg, align_h="center")
        _data_cell(ws, r, 6,  result.matched_count, align_h="center")
        _data_cell(ws, r, 7,  result.total_26as_entries, align_h="center")
        _data_cell(ws, r, 8,  result.unmatched_26as_count, align_h="center")
        _data_cell(ws, r, 9,  result.unmatched_books_count, align_h="center")
        _data_cell(ws, r, 10, result.constraint_violations, bg=viol_bg, align_h="center")
        _data_cell(ws, r, 11, result.high_confidence_count, bg=CONF_HIGH, align_h="center")
        _data_cell(ws, r, 12, result.medium_confidence_count, bg=CONF_MEDIUM, align_h="center")
        _data_cell(ws, r, 13, result.low_confidence_count, bg=CONF_LOW, align_h="center")
        _data_cell(ws, r, 14, result.cross_fy_match_count, align_h="center")
        _data_cell(ws, r, 15, f"{result.avg_variance_pct:.2f}%", align_h="center")
        _data_cell(ws, r, 16, "✓", bg=VAR_GREEN, align_h="center")

        total_26as += result.total_26as_entries
        total_matched += result.matched_count
        total_unmatched_26as += result.unmatched_26as_count
        total_unmatched_books += result.unmatched_books_count
        total_violations += result.constraint_violations
        total_high += result.high_confidence_count
        total_med += result.medium_confidence_count
        total_low += result.low_confidence_count

    # Totals row
    if results:
        tr = len(results) + 3
        _data_cell(ws, tr, 1, "", bold=True)
        _data_cell(ws, tr, 2, "TOTAL", bold=True)
        overall_rate = (total_matched / total_26as * 100) if total_26as > 0 else 0.0
        _data_cell(ws, tr, 5, f"{overall_rate:.1f}%", bold=True, align_h="center")
        _data_cell(ws, tr, 6, total_matched, bold=True, align_h="center")
        _data_cell(ws, tr, 7, total_26as, bold=True, align_h="center")
        _data_cell(ws, tr, 8, total_unmatched_26as, bold=True, align_h="center")
        _data_cell(ws, tr, 9, total_unmatched_books, bold=True, align_h="center")
        _data_cell(ws, tr, 10, total_violations, bold=True, align_h="center")
        _data_cell(ws, tr, 11, total_high, bold=True, align_h="center")
        _data_cell(ws, tr, 12, total_med, bold=True, align_h="center")
        _data_cell(ws, tr, 13, total_low, bold=True, align_h="center")

    _autofit(ws)


def generate_batch_excel(
    results: List[Tuple[RecoResult, CleaningReport, str]],
    fy_label: str = DEFAULT_FINANCIAL_YEAR,
) -> bytes:
    """
    Generate a combined Excel workbook for batch multi-party reconciliation.

    Args:
        results: List of (RecoResult, CleaningReport, sap_filename) tuples
        fy_label: Financial year label

    Returns:
        Excel file bytes
    """
    wb = Workbook()
    used_names: Set[str] = set()

    # Sheet 1: Master Summary
    ws_master = wb.active
    ws_master.title = "Master Summary"
    used_names.add("Master Summary")
    _build_master_summary(ws_master, results, fy_label)

    # Per-party sheets
    for result, report, sap_fn in results:
        party_name = result.deductor_name or result.tan or "Unknown"

        # Matched Pairs sheet
        sheet_name = _safe_sheet_name(party_name, "Match", used_names)
        ws_matched = wb.create_sheet(sheet_name)
        _build_matched(ws_matched, result.matched_pairs)

        # Unmatched 26AS sheet
        sheet_name = _safe_sheet_name(party_name, "Un26AS", used_names)
        ws_un26as = wb.create_sheet(sheet_name)
        _build_unmatched_26as(ws_un26as, result.unmatched_26as)

        # Unmatched Books sheet
        sheet_name = _safe_sheet_name(party_name, "UnBks", used_names)
        ws_unbks = wb.create_sheet(sheet_name)
        _build_unmatched_books(ws_unbks, result.unmatched_books)

        # Variance Analysis sheet
        sheet_name = _safe_sheet_name(party_name, "Var", used_names)
        ws_var = wb.create_sheet(sheet_name)
        _build_variance(ws_var, result.matched_pairs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
