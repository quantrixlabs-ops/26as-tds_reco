"""
Real-data benchmark: runs the reconciliation algorithm against actual Lot 3 SAP files
and the 26AS FY 23-24 master, then compares results against LOT 3 WORKINGS ground truth.

Usage: cd backend && python benchmark_real.py
"""
from __future__ import annotations

import json
import os
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent))

from config import fy_date_range, sap_date_window, ALLOW_CROSS_FY, DEFAULT_FINANCIAL_YEAR
from cleaner import clean_sap_books
from parser_26as import parse_26as
from engine.optimizer import run_global_optimizer, BookEntry, As26Entry
from engine.validator import validate_26as, validate_sap_books, compute_control_totals
from engine.exception_engine import generate_exceptions
from engine.scorer import score_candidate

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE = Path(os.path.expanduser("~/Desktop/Bharath Surya/ABB"))
SAP_DIR = BASE / "Lot 3"
WORKINGS_DIR = BASE / "LOT 3 WORKINGS"
AS26_PATH = BASE / "26AS FY 23-24.xlsx"
FY = "FY2023-24"

# ── Ground truth parser ───────────────────────────────────────────────────────

def parse_workings_ground_truth(filepath: Path) -> dict:
    """Parse a LOT 3 WORKINGS file and extract ground truth metrics."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    result = {
        "matched_count": 0,
        "unmatched_count": 0,
        "unused_books": 0,
        "total_books": 0,
        "total_26as": 0,
        "match_rate": 0.0,
        "matched_amount": 0.0,
        "matched_entries": [],  # list of (26as_amount, books_amount, variance_pct, match_type)
    }

    # Parse Summary sheet
    for shname in wb.sheetnames:
        if shname.lower() == "summary":
            ws = wb[shname]
            for row in ws.iter_rows(values_only=True):
                if row and row[0]:
                    label = str(row[0]).strip().lower()
                    if "total 26as" in label and "entries" in label:
                        try: result["total_26as"] = int(row[1])
                        except: pass
                    elif "matched" in label and "count" in label:
                        try: result["matched_count"] = int(row[1])
                        except: pass
                    elif "unmatched" in label and "count" in label:
                        try: result["unmatched_count"] = int(row[1])
                        except: pass
                    elif "match rate" in label:
                        try: result["match_rate"] = float(str(row[1]).replace("%", ""))
                        except: pass
            break

    # Parse Reconciliation Detail sheet
    for shname in wb.sheetnames:
        if "reconciliation" in shname.lower() or "detail" in shname.lower():
            ws = wb[shname]
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                vals = list(row)
                if not header_found:
                    # Look for header row
                    if any(str(v or "").lower().startswith("26as") for v in vals[:3]):
                        header_found = True
                        continue
                    continue
                # Data row
                try:
                    as26_amt = float(vals[2]) if vals[2] else 0
                    books_amt = float(vals[4]) if vals[4] else 0
                    variance = float(str(vals[6]).replace("%", "")) if vals[6] else 0
                    match_type = str(vals[8]) if len(vals) > 8 and vals[8] else "Unknown"
                    result["matched_entries"].append({
                        "as26_amount": as26_amt,
                        "books_amount": books_amt,
                        "variance_pct": variance,
                        "match_type": match_type,
                    })
                    result["matched_amount"] += as26_amt
                except:
                    pass
            break

    wb.close()
    return result


# ── SAP → BookEntry converter ─────────────────────────────────────────────────

def df_to_book_entries(df: pd.DataFrame) -> list[BookEntry]:
    entries = []
    for idx, row in df.iterrows():
        entries.append(BookEntry(
            index=int(idx),
            invoice_ref=str(row.get("invoice_ref", "")),
            amount=float(row.get("amount", 0)),
            doc_date=str(row.get("doc_date", "")),
            doc_type=str(row.get("doc_type", "")),
            clearing_doc=str(row.get("clearing_doc", "")),
            sap_fy=str(row.get("sap_fy", "")),
            flag=str(row.get("flag", "")),
        ))
    return entries


def df_to_as26_entries(df: pd.DataFrame) -> list[As26Entry]:
    entries = []
    for idx, row in df.iterrows():
        entries.append(As26Entry(
            index=int(idx),
            amount=float(row.get("amount", 0)),
            transaction_date=str(row.get("transaction_date", "")),
            section=str(row.get("section", "")),
            tan=str(row.get("tan", "")),
            deductor_name=str(row.get("deductor_name", "")),
        ))
    return entries


# ── Main benchmark ────────────────────────────────────────────────────────────

def run_single_benchmark(sap_name: str) -> dict:
    """Run reconciliation for a single deductor and return metrics."""
    sap_path = SAP_DIR / sap_name
    sap_bytes = sap_path.read_bytes()
    as26_bytes = AS26_PATH.read_bytes()

    fy_start, fy_end = fy_date_range(FY)
    sap_start, sap_end = sap_date_window(FY)

    # Clean SAP
    t0 = time.time()
    clean_df, _sgl_v_df, cleaning_report = clean_sap_books(sap_bytes, fy_start=sap_start, fy_end=sap_end)
    t_clean = time.time() - t0

    # Parse 26AS
    t0 = time.time()
    as26_df = parse_26as(as26_bytes, fy_start=fy_start, fy_end=fy_end)
    t_parse = time.time() - t0

    # Fuzzy match deductor name from SAP filename
    deductor_name = sap_name.replace(".XLSX", "").replace(".xlsx", "").replace("_", " ").strip()

    # Filter 26AS to this deductor (fuzzy match on name)
    from rapidfuzz import fuzz
    scores = as26_df.groupby("deductor_name").size().reset_index(name="count")
    scores["score"] = scores["deductor_name"].apply(
        lambda x: fuzz.token_sort_ratio(deductor_name.upper(), x.upper())
    )
    best_matches = scores[scores["score"] >= 70].sort_values("score", ascending=False)

    if best_matches.empty:
        return {
            "deductor": deductor_name,
            "status": "NO_MATCH",
            "error": f"No 26AS deductor matched '{deductor_name}' (best < 70)",
        }

    # Take top match(es)
    matched_names = best_matches.head(3)["deductor_name"].tolist()
    filtered_26as = as26_df[as26_df["deductor_name"].isin(matched_names)].copy()

    if filtered_26as.empty:
        return {
            "deductor": deductor_name,
            "status": "NO_26AS_ENTRIES",
            "matched_deductor": matched_names[0] if matched_names else None,
        }

    # Validate
    validated_df, val_report = validate_26as(filtered_26as)
    clean_df, sap_issues = validate_sap_books(clean_df)

    # Build entries
    book_entries = df_to_book_entries(clean_df)
    valid_26as = validated_df[validated_df["_valid"] == True]
    as26_entries = df_to_as26_entries(valid_26as)

    current_books = [b for b in book_entries if b.sap_fy == FY or not b.sap_fy]
    prior_books = [b for b in book_entries if b.sap_fy and b.sap_fy != FY]

    # Run optimizer
    t0 = time.time()
    all_results, unmatched_entries = run_global_optimizer(
        as26_entries=as26_entries,
        book_pool=book_entries,
        current_books=current_books,
        prior_books=prior_books,
        allow_cross_fy=ALLOW_CROSS_FY,
    )
    t_opt = time.time() - t0
    matched_results = [r for r in all_results if not r.suggested]
    suggested_results = [r for r in all_results if r.suggested]

    # Metrics
    total_26as = len(as26_entries)
    matched_count = len(matched_results)
    unmatched_count = len(unmatched_entries)
    match_rate = (matched_count / total_26as * 100) if total_26as > 0 else 0

    high_conf = sum(1 for r in matched_results if r.confidence == "HIGH")
    med_conf = sum(1 for r in matched_results if r.confidence == "MEDIUM")
    low_conf = sum(1 for r in matched_results if r.confidence == "LOW")

    # Compliance checks — track by unique book index, not invoice_ref
    invoice_reuse = 0
    seen_indices = set()
    for r in matched_results:
        for b in r.books:
            if b.index in seen_indices:
                invoice_reuse += 1
            seen_indices.add(b.index)

    books_sum_fn = lambda r: sum(b.amount for b in r.books)
    books_exceed = sum(1 for r in matched_results if books_sum_fn(r) > r.as26_amount + 0.01)

    # Match type distribution
    type_dist = {}
    for r in matched_results:
        t = r.match_type
        type_dist[t] = type_dist.get(t, 0) + 1

    # Variance distribution
    var_buckets = {"<1%": 0, "1-2%": 0, "2-3%": 0, "3-5%": 0, ">5%": 0}
    for r in matched_results:
        v = r.variance_pct
        if v < 1:
            var_buckets["<1%"] += 1
        elif v < 2:
            var_buckets["1-2%"] += 1
        elif v < 3:
            var_buckets["2-3%"] += 1
        elif v < 5:
            var_buckets["3-5%"] += 1
        else:
            var_buckets[">5%"] += 1

    return {
        "deductor": deductor_name,
        "matched_deductor_26as": matched_names[0],
        "fuzzy_score": int(best_matches.iloc[0]["score"]),
        "status": "OK",
        "sap_rows_raw": len(clean_df) + cleaning_report.excluded_total if hasattr(cleaning_report, 'excluded_total') else len(clean_df),
        "sap_rows_clean": len(clean_df),
        "as26_entries": total_26as,
        "matched": matched_count,
        "unmatched": unmatched_count,
        "match_rate_pct": round(match_rate, 2),
        "high_confidence": high_conf,
        "medium_confidence": med_conf,
        "low_confidence": low_conf,
        "match_types": type_dist,
        "variance_dist": var_buckets,
        "invoice_reuse_violations": invoice_reuse,
        "books_exceed_violations": books_exceed,
        "current_fy_books": len(current_books),
        "prior_fy_books": len(prior_books),
        "time_clean_s": round(t_clean, 2),
        "time_parse_s": round(t_parse, 2),
        "time_optimize_s": round(t_opt, 2),
        "time_total_s": round(t_clean + t_parse + t_opt, 2),
    }


def compare_with_ground_truth(algo_result: dict, gt: dict) -> dict:
    """Compare algorithm result against LOT 3 WORKINGS ground truth."""
    if algo_result["status"] != "OK":
        return {"comparison": "SKIPPED", "reason": algo_result.get("error", algo_result["status"])}

    algo_matched = algo_result["matched"]
    gt_matched = gt["matched_count"]
    algo_rate = algo_result["match_rate_pct"]
    gt_rate = gt["match_rate"]

    return {
        "algo_matched": algo_matched,
        "gt_matched": gt_matched,
        "diff_matched": algo_matched - gt_matched,
        "algo_rate": algo_rate,
        "gt_rate": gt_rate,
        "diff_rate": round(algo_rate - gt_rate, 2),
        "algo_better": algo_matched >= gt_matched,
        "compliance_clean": algo_result["invoice_reuse_violations"] == 0 and algo_result["books_exceed_violations"] == 0,
    }


def main():
    print("=" * 80)
    print("REAL DATA BENCHMARK — LOT 3 vs 26AS FY 23-24")
    print("=" * 80)

    # Verify files exist
    if not AS26_PATH.exists():
        print(f"ERROR: 26AS file not found at {AS26_PATH}")
        return
    if not SAP_DIR.exists():
        print(f"ERROR: SAP dir not found at {SAP_DIR}")
        return

    # Get SAP files (exclude temp files and customer codes)
    sap_files = sorted([
        f for f in os.listdir(SAP_DIR)
        if f.endswith((".XLSX", ".xlsx")) and not f.startswith("~$") and "Customer Codes" not in f
    ])
    print(f"\nFound {len(sap_files)} SAP files")

    # Get ground truth files
    gt_map = {}
    if WORKINGS_DIR.exists():
        for f in os.listdir(WORKINGS_DIR):
            if f.endswith((".xlsx", ".XLSX")) and not f.startswith("~$"):
                # Extract deductor name from "1 VASHI ELECTRICALS PVT LTD M S.xlsx"
                name = f.rsplit(".", 1)[0]
                # Remove leading number
                parts = name.split(" ", 1)
                if len(parts) > 1 and parts[0].isdigit():
                    name = parts[1]
                gt_map[name.upper().strip()] = WORKINGS_DIR / f
        print(f"Found {len(gt_map)} ground truth WORKINGS files")

    # Pre-parse 26AS once (it's 61K rows)
    print("\nPre-loading 26AS FY 23-24...")
    t0 = time.time()
    as26_bytes_cached = AS26_PATH.read_bytes()
    print(f"  26AS loaded in {time.time() - t0:.1f}s ({len(as26_bytes_cached) / 1024 / 1024:.1f} MB)")

    results = []
    comparisons = []
    total_start = time.time()

    for i, sap_file in enumerate(sap_files):
        deductor = sap_file.replace(".XLSX", "").replace(".xlsx", "")
        print(f"\n[{i+1}/{len(sap_files)}] {deductor}")
        print("-" * 60)

        try:
            result = run_single_benchmark(sap_file)
            results.append(result)

            if result["status"] == "OK":
                print(f"  26AS match: {result['matched_deductor_26as']} (score={result['fuzzy_score']})")
                print(f"  SAP: {result['sap_rows_clean']} clean rows | 26AS: {result['as26_entries']} entries")
                print(f"  Matched: {result['matched']}/{result['as26_entries']} ({result['match_rate_pct']}%)")
                print(f"  Confidence: HIGH={result['high_confidence']} MED={result['medium_confidence']} LOW={result['low_confidence']}")
                print(f"  Types: {result['match_types']}")
                print(f"  Variance: {result['variance_dist']}")
                print(f"  Compliance: reuse={result['invoice_reuse_violations']} exceed={result['books_exceed_violations']}")
                print(f"  Time: {result['time_total_s']}s (clean={result['time_clean_s']}s parse={result['time_parse_s']}s opt={result['time_optimize_s']}s)")

                # Compare with ground truth
                gt_key = deductor.upper().strip()
                gt_found = None
                from rapidfuzz import fuzz as fz
                for gk, gv in gt_map.items():
                    if fz.token_sort_ratio(gt_key, gk) >= 80:
                        gt_found = (gk, gv)
                        break

                if gt_found:
                    gt_data = parse_workings_ground_truth(gt_found[1])
                    comp = compare_with_ground_truth(result, gt_data)
                    comp["deductor"] = deductor
                    comparisons.append(comp)
                    marker = ">>>" if comp.get("algo_better") else "<<<"
                    print(f"  {marker} vs Ground Truth: algo={comp['algo_matched']} gt={comp['gt_matched']} (diff={comp['diff_matched']:+d})")
                else:
                    print(f"  (no ground truth found)")
            else:
                print(f"  SKIPPED: {result.get('error', result['status'])}")

        except Exception as e:
            print(f"  ERROR: {e}")
            traceback.print_exc()
            results.append({"deductor": deductor, "status": "ERROR", "error": str(e)})

    # ── Summary ────────────────────────────────────────────────────────────────
    total_time = time.time() - total_start
    ok_results = [r for r in results if r["status"] == "OK"]

    print("\n" + "=" * 80)
    print("BENCHMARK SUMMARY")
    print("=" * 80)
    print(f"Total deductors: {len(results)}")
    print(f"Successful: {len(ok_results)}")
    print(f"Failed/Skipped: {len(results) - len(ok_results)}")
    print(f"Total time: {total_time:.1f}s")

    if ok_results:
        total_26as = sum(r["as26_entries"] for r in ok_results)
        total_matched = sum(r["matched"] for r in ok_results)
        total_unmatched = sum(r["unmatched"] for r in ok_results)
        avg_rate = sum(r["match_rate_pct"] for r in ok_results) / len(ok_results)
        total_reuse = sum(r["invoice_reuse_violations"] for r in ok_results)
        total_exceed = sum(r["books_exceed_violations"] for r in ok_results)

        print(f"\n  Total 26AS entries processed: {total_26as}")
        print(f"  Total matched: {total_matched}")
        print(f"  Total unmatched: {total_unmatched}")
        print(f"  Average match rate: {avg_rate:.1f}%")
        print(f"  Invoice reuse violations: {total_reuse}")
        print(f"  Books > 26AS violations: {total_exceed}")

        # Aggregate match types
        all_types = {}
        for r in ok_results:
            for k, v in r["match_types"].items():
                all_types[k] = all_types.get(k, 0) + v
        print(f"\n  Match type distribution:")
        for k, v in sorted(all_types.items(), key=lambda x: -x[1]):
            print(f"    {k}: {v} ({v/total_matched*100:.1f}%)")

        # Aggregate variance
        all_var = {"<1%": 0, "1-2%": 0, "2-3%": 0, "3-5%": 0, ">5%": 0}
        for r in ok_results:
            for k, v in r["variance_dist"].items():
                all_var[k] = all_var.get(k, 0) + v
        print(f"\n  Variance distribution:")
        for k, v in all_var.items():
            print(f"    {k}: {v} ({v/total_matched*100:.1f}%)" if total_matched > 0 else f"    {k}: {v}")

        # Confidence breakdown
        total_high = sum(r["high_confidence"] for r in ok_results)
        total_med = sum(r["medium_confidence"] for r in ok_results)
        total_low = sum(r["low_confidence"] for r in ok_results)
        if total_matched > 0:
            print(f"\n  Confidence: HIGH={total_high} ({total_high/total_matched*100:.1f}%) MED={total_med} ({total_med/total_matched*100:.1f}%) LOW={total_low} ({total_low/total_matched*100:.1f}%)")
        else:
            print(f"\n  Confidence: HIGH={total_high} MED={total_med} LOW={total_low}")

    if comparisons:
        valid_comps = [c for c in comparisons if "algo_matched" in c]
        algo_wins = sum(1 for c in valid_comps if c.get("algo_better"))
        print(f"\n  Ground Truth Comparison ({len(valid_comps)} deductors):")
        print(f"    Algorithm better/equal: {algo_wins}/{len(valid_comps)}")
        total_algo = sum(c["algo_matched"] for c in valid_comps)
        total_gt = sum(c["gt_matched"] for c in valid_comps)
        print(f"    Total algo matched: {total_algo} vs GT: {total_gt} (diff: {total_algo - total_gt:+d})")
        compliance_clean = sum(1 for c in valid_comps if c.get("compliance_clean"))
        print(f"    Compliance clean: {compliance_clean}/{len(valid_comps)}")

    # Save results to JSON
    output_path = Path(__file__).parent / "benchmark_results.json"
    with open(output_path, "w") as f:
        json.dump({
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "financial_year": FY,
            "total_deductors": len(results),
            "successful": len(ok_results),
            "total_time_s": round(total_time, 1),
            "results": results,
            "comparisons": comparisons,
        }, f, indent=2, default=str)
    print(f"\nResults saved to {output_path}")


if __name__ == "__main__":
    main()
